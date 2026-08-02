"""
recorder.py — Excel 持仓记录读写

功能：
1. add_purchase()    — 追加一条买入记录到 portfolio.xlsx
2. get_all_records() — 读取所有历史记录，返回 list[dict]

表头（Sheet1）：
    日期 | 标的代码 | 标的名称 | 买入价格 | 买入份额 | 实际花费 | 手续费 | 累计投入 | 累计份额 | 备注
"""

import os
from datetime import date
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------- 配置 ----------------------
from config import ETF_CODE, ETF_NAME, EXCEL_PATH

# ---------------------- 表头定义 ----------------------
HEADERS = [
    "日期",
    "标的代码",
    "标的名称",
    "买入价格",
    "买入份额",
    "实际花费",
    "手续费",
    "累计投入",
    "累计份额",
    "累计滚存",
    "备注",
]

# 表头样式
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")  # 蓝色底
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

# 边框
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _load_workbook(path: str):
    """加载或新建 Workbook，自动创建 Sheet1。"""
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        if "Sheet1" not in wb.sheetnames:
            ws = wb.create_sheet("Sheet1")
        else:
            ws = wb["Sheet1"]
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        _write_headers(ws)
        wb.save(path)
    return wb, ws


def _write_headers(ws):
    """写入表头（已有内容时跳过），并设置列宽。"""
    # 如果已有数据行（或表头已存在），跳过写入
    if ws.max_row >= 1 and ws.cell(1, 1).value is not None:
        # 表头已存在，只确保列宽
        col_widths = [14, 12, 18, 12, 12, 12, 10, 12, 12, 12, 20]
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.freeze_panes = "A2"
        return
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = CELL_BORDER

    # 列宽（手动微调）
    col_widths = [14, 12, 18, 12, 12, 12, 10, 12, 12, 12, 20]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # 首行冻结


def _apply_row_style(ws, row_idx: int):
    """给数据行加上边框和交替底色。"""
    fill_color = "DEEAF1" if row_idx % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(fill_type="solid", fgColor=fill_color)
    for col_idx in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx not in (3, 10):  # 名称和备注不居中
            cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = row_fill


# ====================== 公开函数 ======================


def add_purchase(
    buy_date: str,
    buy_price: float,
    shares: int,
    cost: float,
    fee: float,
    total_invest: float,
    total_shares: int,
    rollover: float = 0.0,
    remark: Optional[str] = None,
) -> dict:
    """
    追加一条买入记录到 portfolio.xlsx。

    参数：
        buy_date      : str  — 买入日期，格式 YYYY-MM-DD
        buy_price     : float — 成交单价（元/份）
        shares        : int   — 买入份额（份）
        cost          : float — 实际花费（元）
        fee           : float — 手续费（元）
        total_invest  : float — 累计投入（元）
        total_shares  : int   — 累计份额（份）
        rollover      : float — 本期滚存余额（元）
        remark        : str   — 备注（可选）

    返回：
        dict — 本条记录内容（用于确认）
    """
    record = {
        "日期":       buy_date,
        "标的代码":   ETF_CODE,
        "标的名称":   ETF_NAME,
        "买入价格":   buy_price,
        "买入份额":   shares,
        "实际花费":   cost,
        "手续费":     fee,
        "累计投入":   total_invest,
        "累计份额":   total_shares,
        "累计滚存":   round(rollover, 2),
        "备注":       remark or "",
    }

    wb, ws = _load_workbook(EXCEL_PATH)

    # 确保表头存在（新建 Sheet 时）
    _write_headers(ws)

    # 追加到最新一行的下一行
    next_row = ws.max_row + 1
    for col_idx, key in enumerate(HEADERS, start=1):
        ws.cell(row=next_row, column=col_idx, value=record[key])

    _apply_row_style(ws, next_row)

    wb.save(EXCEL_PATH)

    print(f"[记录已保存] 第{next_row - 1}条 | {buy_date} | "
          f"{shares}份 @{buy_price:.3f} | 累计投入 ¥{total_invest:.2f}")

    # 止盈重置：如果上一轮止盈已结束(state==3)，新买入触发重置
    try:
        import profit_taker
        st = profit_taker._load_state()
        if st["state"] == 3:
            profit_taker.reset()
    except Exception:
        pass

    return record


def get_all_records() -> list[dict]:
    """
    读取 portfolio.xlsx Sheet1 的所有数据行，返回 list[dict]。

    返回：
        list[dict] — 每条记录为一个 dict，按原始顺序排列。
                     空文件则返回空列表 []。
    """
    if not os.path.exists(EXCEL_PATH):
        return []

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) <= 1:  # 只有表头或空文件
        return []

    header_row = rows[0]
    records = []
    for row in rows[1:]:
        if row[0] is None:  # 跳过空行
            continue
        record = dict(zip(header_row, row))
        records.append(record)

    return records


def print_records():
    """打印所有记录（供命令行查看）。"""
    records = get_all_records()
    if not records:
        print("[记录为空] 请先运行 add_purchase() 添加数据")
        return

    print(f"\n{'=' * 90}")
    print(f"{'序号':^4} {'日期':^12} {'价格':^8} {'份额':^8} {'花费':^10} "
          f"{'手续费':^8} {'累计投入':^10} {'累计份额':^10} {'备注'}")
    print(f"{'=' * 90}")
    for i, r in enumerate(records, start=1):
        print(f"{i:^4} {str(r['日期']):^12} {r['买入价格']:^8.3f} "
              f"{r['买入份额']:^8} {r['实际花费']:^10.2f} {r['手续费']:^8.2f} "
              f"{r['累计投入']:^10.2f} {r['累计份额']:^10} {r.get('备注', '')}")
    print(f"{'=' * 90}")
    print(f"共 {len(records)} 条记录")


# ====================== 测试 ======================
if __name__ == "__main__":
    # 示例：写入一条测试记录
    add_purchase(
        buy_date="2026-07-20",
        buy_price=3.924,
        shares=100,
        cost=392.40,
        fee=5.70,
        total_invest=398.10,
        total_shares=100,
        remark="Day3 第一次真实买入",
    )

    # 再追加一条
    add_purchase(
        buy_date="2026-07-21",
        buy_price=3.885,
        shares=100,
        cost=388.50,
        fee=5.00,
        total_invest=792.00,
        total_shares=200,
        remark="弹性股数法，滚存¥111.50",
    )

    print_records()
