"""
inspector.py — 每日巡检纯函数（Day11 改版）
============================================
纯函数，无 input()，无业务副作用（不改 xlsx）。
check() 更新 highest_price 属"观测追踪"，非业务动作。
可被 Cron / Skill / main 任意调用。

五个检查点：
  B-1 准备：返回 ok（不需检查）
  B-2 建仓：已有 Day7 Cron，跳过
  B-3 持有 PE：>70% 提醒偏高，<20% 方案A加仓，<30% 推送
  B-4 止盈：调 profit_taker.check()，返回触发状态
  B-5 纪律：浮亏 >= 20% 推哨点 + 别跳投检查

推送逻辑（should_push）：
  浮盈>=20% / 浮亏>=15% / PE<30% / PE>70% / 止盈触发 /
  下次定投<=2天 / 别跳投 → 推送；其余静默。
"""
import os
import sys
from datetime import datetime, date

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

TOOLS_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(TOOLS_DIR)
sys.path.insert(0, TOOLS_DIR)

# 路径修正
import recorder
recorder.EXCEL_PATH = os.path.join(PROJECT_DIR, "data", "portfolio.xlsx")

EXCEL_FILE = os.path.join(PROJECT_DIR, "data", "portfolio.xlsx")


def _check_skip() -> str:
    """
    别跳投检查：定投日已过但无对应买入记录 -> warn_skip。
    读 Sheet3 计划 + Sheet1 记录，纯读取不写。
    """
    try:
        import openpyxl
        if not os.path.exists(EXCEL_FILE):
            return "ok"
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)

        # 读 Sheet3 计划
        if "Sheet3" not in wb.sheetnames:
            wb.close()
            return "ok"
        ws3 = wb["Sheet3"]
        today = date.today()
        overdue_periods = []
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            no = int(row[0])
            planned_str = str(row[1]) if row[1] else ""
            status = str(row[4]) if row[4] else ""
            if "已完成" in status:
                continue
            try:
                planned_date = datetime.strptime(planned_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            if planned_date < today:
                days_overdue = (today - planned_date).days
                overdue_periods.append((no, planned_str, days_overdue))

        wb.close()

        if overdue_periods:
            no, planned, days = overdue_periods[0]  # 报最近的一期
            return (f"warn_skip: 第{no}期(计划{planned})已过{days}天未买入，"
                    f"别跳投、按计划补齐")
        return "ok"
    except Exception:
        return "ok"


def _check_next_dca() -> tuple:
    """
    查 Sheet3 找下一期未完成的定投，返回 (期数, 计划日期str, 剩余天数)。
    无待执行期 → (0, '', 999)。
    """
    try:
        import openpyxl
        if not os.path.exists(EXCEL_FILE):
            return (0, "", 999)
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
        if "Sheet3" not in wb.sheetnames:
            wb.close()
            return (0, "", 999)
        ws3 = wb["Sheet3"]
        today = date.today()
        best = (0, "", 999)
        for row in ws3.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            status = str(row[4]) if row[4] else ""
            if "已完成" in status:
                continue
            no = int(row[0])
            planned_str = str(row[1]) if row[1] else ""
            try:
                planned_date = datetime.strptime(planned_str, "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            days_left = (planned_date - today).days
            if days_left < best[2]:
                best = (no, planned_str, days_left)
        wb.close()
        return best
    except Exception:
        return (0, "", 999)


def inspect_once() -> dict:
    """
    执行全部巡检，返回结构化结论。

    返回 dict:
        price          : float  — 当前价
        pnl_pct        : float  — 浮盈亏比(%)
        profit_action  : str    — 止盈判断结论
        profit_trigger : bool   — 是否触发止盈
        pe_pct         : float  — PE 分位(%)
        pe_alert       : str    — PE 结论
        warning_20     : bool   — 浮亏>=20%?
        skip_alert     : str    — 别跳投结论
        should_push    : bool   — 是否应推送（True=有情况，False=静默）
        push_reasons   : list   — 推送原因列表（空=无情况）
        summary        : str    — 一段话总结（可直接推微信）
    """
    from portfolio import get_price_now, get_pe_data
    import profit_taker

    result = {
        "price": 0,
        "pnl_pct": 0,
        "profit_action": "ok",
        "profit_trigger": False,
        "pe_pct": 0,
        "pe_alert": "ok",
        "warning_20": False,
        "skip_alert": "ok",
        "should_push": False,
        "push_reasons": [],
        "summary": "",
    }

    # B-1 准备：不需定时检查
    # B-2 建仓：已有 Day7 双周 Cron

    # 获取当前价格
    price = get_price_now()
    if price:
        result["price"] = price

    # B-3 持有 PE
    try:
        pe_data = get_pe_data()
        if pe_data and "pe_5y_pct" in pe_data:
            pe_pct = pe_data["pe_5y_pct"]
            result["pe_pct"] = pe_pct
            if pe_pct > 70:
                result["pe_alert"] = f"pe_high: 估值偏高({pe_pct:.1f}%)，仅提醒不触发卖出"
            elif pe_pct < 20:
                # 方案A（等深坑）：PE极度低估时加仓
                # 货基余额 = 历史止盈卖出金额之和
                try:
                    import profit_taker
                    pt_state = profit_taker._load_state()
                    mf_balance = sum(r.get("amount", 0)
                                     for r in pt_state.get("sell_records", []))
                except Exception:
                    mf_balance = 0
                add_limit = min(500, mf_balance * 0.5)
                add_amount = min(add_limit, mf_balance)
                add_amount = round(add_amount, 2)
                result["pe_alert"] = (
                    f"pe_low: 市场极度低估（PE分位{pe_pct:.1f}% < 20%）。"
                    f"你有 ¥{mf_balance:.0f} 在货基里，"
                    f"可考虑额外加仓 ¥{add_amount:.0f}"
                    f"（不超过 ¥500，且不超过货基的 50%）。"
                    f"是否加仓、由你判断。"
                )
                result["mf_balance"] = mf_balance
                result["pe_add_amount"] = add_amount
            else:
                result["pe_alert"] = f"pe_normal: 估值中性({pe_pct:.1f}%)"
    except Exception:
        result["pe_alert"] = "pe_error: PE数据获取失败"

    # B-4 止盈
    if price and price > 0:
        try:
            pt_result = profit_taker.check(price)
            result["profit_action"] = pt_result["action"]
            result["profit_trigger"] = pt_result["trigger"]
            if pt_result["trigger"]:
                result["profit_action"] = (
                    f"sell_tier: {pt_result['action']}，"
                    f"建议卖 {pt_result['sell_shares']:.0f} 份"
                )
        except Exception:
            result["profit_action"] = "profit_error: 止盈检查失败"

    # B-5 纪律（浮亏哨点）
    try:
        records = recorder.get_all_records()
        if records and price and price > 0:
            last = records[-1]
            total_invest = float(last["累计投入"])
            total_shares = float(last["累计份额"])
            if total_shares > 0 and total_invest > 0:
                avg_cost = total_invest / total_shares
                pnl_pct = (price - avg_cost) / avg_cost * 100
                result["pnl_pct"] = round(pnl_pct, 2)
                if pnl_pct <= -20:
                    result["warning_20"] = True
    except Exception:
        pass

    # B-5 纪律（别跳投）
    result["skip_alert"] = _check_skip()

    # ── 推送判断（每日轻量，无情况静默）──
    push_reasons = []

    # 1) 浮盈 ≥ 20%：接近止盈线
    if result["pnl_pct"] >= 20:
        push_reasons.append(f"浮盈{result['pnl_pct']:.1f}%，接近止盈线")

    # 2) 浮亏 ≥ 15%：回撤预警（20%哨点前哨）
    if result["pnl_pct"] <= -15:
        push_reasons.append(f"浮亏{result['pnl_pct']:.1f}%，回撤预警")

    # 3) PE < 30%：低估值，考虑货基加回
    if result["pe_pct"] > 0 and result["pe_pct"] < 30:
        push_reasons.append(f"PE分位{result['pe_pct']:.1f}%<30%，低估值")

    # 4) PE > 70%：高估值，关注止盈
    if result["pe_pct"] > 70:
        push_reasons.append(f"PE分位{result['pe_pct']:.1f}%>70%，高估值")

    # 5) 止盈触发
    if result["profit_trigger"]:
        push_reasons.append(f"止盈触发：{result['profit_action']}")

    # 6) 下次定投 ≤ 2天
    dca_no, dca_date, dca_days = _check_next_dca()
    if 0 <= dca_days <= 2:
        push_reasons.append(f"第{dca_no}期定投{dca_date}到期(还剩{dca_days}天)")

    # 7) 别跳投
    if result["skip_alert"] != "ok":
        push_reasons.append(result["skip_alert"])

    result["push_reasons"] = push_reasons
    result["should_push"] = len(push_reasons) > 0

    # 生成 summary（仅 should_push 时有实质内容）
    if result["should_push"]:
        parts = []
        parts.append(f"巡检：当前价 {result['price']:.3f} 元")
        parts.append(f"浮盈亏 {result['pnl_pct']:.2f}%")
        if result["pe_alert"] != "ok":
            parts.append(result["pe_alert"])
        if result["profit_trigger"]:
            parts.append(f"[触发] {result['profit_action']}")
        if result["warning_20"]:
            parts.append("[预警] 浮亏超20%，请核对：闲钱?现金流?纪律?——不卖，只看")
        if result["skip_alert"] != "ok":
            parts.append(f"[纪律] {result['skip_alert']}")
        parts.append("原因：" + "；".join(push_reasons))
        result["summary"] = "。".join(parts) + "。不替你下单，只给结论。"
    else:
        result["summary"] = "一切正常，无需操作。"

    return result


# 兼容旧名
check_all = inspect_once


if __name__ == "__main__":
    r = inspect_once()
    print(r["summary"])
