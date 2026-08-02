"""
scheduler.py — 定投日历生成与状态管理

功能：
1. generate_schedule(first_date, count)
   — 从 first_date 开始，每14天一次，生成 count 期计划列表
   
2. init_schedule(first_date, count)
   — 生成全部计划，写入 portfolio.xlsx Sheet3（不存在则创建）
   
3. mark_done(no, actual_date)
   — 把第 no 期的 actual_date 和 status 更新为"已完成 ✅"
"""

import os
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 模块级路径（main.py 会 patch 为绝对路径）
EXCEL_PATH = "portfolio.xlsx"

# ==================== 样式 ====================
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

FONT_TITLE = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
FILL_DONE   = PatternFill(fill_type="solid", fgColor="C6EFCE")
FILL_TODAY  = PatternFill(fill_type="solid", fgColor="FFEB9C")
FILL_FUTURE = PatternFill(fill_type="solid", fgColor="DDEBF7")
FILL_HEADER = PatternFill(fill_type="solid", fgColor="4472C4")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

SHEET3_HEADERS = ["期数", "计划日期", "星期", "计划金额", "状态", "实际日期", "实际价格", "实际份额", "备注"]


# ==================== Sheet3 工具 ====================


def _get_or_create_sheet3(wb: openpyxl.Workbook):
    """确保 Sheet3 存在并写入表头。"""
    if "Sheet3" in wb.sheetnames:
        return wb["Sheet3"]
    ws = wb.create_sheet("Sheet3")
    for col_idx, h in enumerate(SHEET3_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = FONT_TITLE
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = CELL_BORDER
    widths = [6, 14, 6, 10, 10, 14, 12, 10, 22]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"
    return ws


# ==================== 1. 生成日历 ====================


def generate_schedule(first_date: str, count: int) -> list[dict]:
    """
    从 first_date 开始，每14天生成 count 期定投计划。

    参数：
        first_date : str  — 第一期日期，格式 "YYYY-MM-DD"
        count     : int  — 总期数

    返回：
        list[dict]，每项：
            no      : int    — 期数（1起）
            planned : str    — 计划日期 "YYYY-MM-DD"
            weekday : str    — 星期几（如"周一"）
            amount  : float  — 计划金额
            status  : str    — 状态标记
            actual  : str    — 实际日期（空字符串表示未执行）
    """
    from config import DCA_AMOUNT, DCA_INTERVAL_DAYS

    start = datetime.strptime(first_date, "%Y-%m-%d")
    weekday_map = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

    schedule = []
    for i in range(1, count + 1):
        plan_date = start + timedelta(days=(i - 1) * DCA_INTERVAL_DAYS)
        schedule.append({
            "no":      i,
            "planned": plan_date.strftime("%Y-%m-%d"),
            "weekday": weekday_map[plan_date.weekday()],
            "amount":  DCA_AMOUNT,
            "status":  "⏳待执行",
            "actual":  "",
        })
    return schedule


# ==================== 2. 写入 Excel ====================


def init_schedule(first_date: str, count: int, excel_path: str = "portfolio.xlsx") -> None:
    """
    生成全部定投计划，写入 portfolio.xlsx Sheet3。
    如果 Sheet3 已有数据则跳过（不重复写入）。
    """
    from config import DCA_AMOUNT

    # 检查是否已有数据
    if os.path.exists(excel_path):
        wb_check = openpyxl.load_workbook(excel_path)
        if "Sheet3" in wb_check.sheetnames:
            ws_existing = wb_check["Sheet3"]
            if ws_existing.max_row > 1:
                print(f"[日历] Sheet3 已有 {ws_existing.max_row - 1} 期计划，跳过初始化")
                wb_check.close()
                return
        wb_check.close()

    schedule = generate_schedule(first_date, count)

    wb = openpyxl.load_workbook(excel_path)
    ws = _get_or_create_sheet3(wb)

    for item in schedule:
        next_row = ws.max_row + 1
        row_data = [
            item["no"], item["planned"], item["weekday"],
            item["amount"], item["status"], item["actual"],
            "", "", "",
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.border = CELL_BORDER
            cell.alignment = ALIGN_CENTER
            cell.fill = FILL_FUTURE

    wb.save(excel_path)
    wb.close()
    print(f"[日历] 已写入 {count} 期定投计划到 portfolio.xlsx Sheet3")


# ==================== 3. 标记完成 ====================


def mark_done(no: int, actual_date: str,
              actual_price: float = 0,
              actual_shares: int = 0,
              remark: str = "",
              excel_path: str = "portfolio.xlsx") -> None:
    """
    把第 no 期标记为"已完成 ✅"，并填写实际执行信息。
    """
    if not os.path.exists(excel_path):
        print("[错误] portfolio.xlsx 不存在，请先执行 add_purchase()")
        return

    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Sheet3"] if "Sheet3" in wb.sheetnames else None
    if ws is None:
        print("[错误] Sheet3 不存在，请先运行 init_schedule()")
        wb.close()
        return

    target_row = None
    for row in ws.iter_rows(min_row=2):
        if row[0].value == no:
            target_row = row[0].row
            break

    if target_row is None:
        print(f"[错误] 未找到第 {no} 期计划")
        wb.close()
        return

    ws.cell(row=target_row, column=5, value="✅已完成")
    ws.cell(row=target_row, column=6, value=actual_date)
    ws.cell(row=target_row, column=7, value=actual_price)
    ws.cell(row=target_row, column=8, value=actual_shares)
    if remark:
        ws.cell(row=target_row, column=9, value=remark)

    for col in range(1, 10):
        ws.cell(row=target_row, column=col).fill = FILL_DONE

    wb.save(excel_path)
    wb.close()

    plan_date = ws.cell(row=target_row, column=2).value
    print(f"[日历] 第{no}期 ✅ | 计划{plan_date} | 实际{actual_date} "
          f"| ¥{actual_price} x {actual_shares}份")


# ==================== 打印日历 ====================


def print_schedule(first_date: str, count: int) -> None:
    """打印完整定投日历到控制台（读取 Sheet3 实际完成状态）。"""
    schedule = generate_schedule(first_date, count)

    # 尝试从 Sheet3 读取实际状态
    done_map = {}  # {期数: 状态字符串}
    if os.path.exists(EXCEL_PATH):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            if "Sheet3" in wb.sheetnames:
                ws = wb["Sheet3"]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] is not None:
                        no = int(row[0])
                        status = str(row[4]) if row[4] else "⏳待执行"
                        done_map[no] = status
            wb.close()
        except Exception:
            pass

    print("\n" + "=" * 62)
    print(f"  双周定投日历（共 {count} 期）")
    print(f"  第1期：{schedule[0]['planned']}  |  第{count}期：{schedule[count-1]['planned']}")
    print("=" * 62)
    print(f"  {'期':^3}  {'日期':^12}  {'星期':^4}  {'金额':^8}  状态")
    print("-" * 62)
    for item in schedule:
        status = done_map.get(item['no'], item['status'])
        print(f"  {item['no']:^3}.  {item['planned']}  {item['weekday']:^4}  "
              f"¥{item['amount']:>6.0f}    {status}")
    print("=" * 62)


# ==================== 测试 ====================
if __name__ == "__main__":
    from config import FIRST_DATE, TOTAL_COUNT
    print_schedule(FIRST_DATE, TOTAL_COUNT)
    print(f"\n验证：第1期={FIRST_DATE}，第{TOTAL_COUNT}期="
          f"{(datetime.strptime(FIRST_DATE,'%Y-%m-%d') + timedelta(days=(TOTAL_COUNT-1)*14)).strftime('%Y-%m-%d')}")
